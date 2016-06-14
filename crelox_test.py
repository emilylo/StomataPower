# Plots cotyledon outlines as polygons and tests the autocorrelation function of
# stomatal positions within the polygons versus that of random points within
# the polygons.

import sys
import numpy as np;
import openpyxl;
import matplotlib.mlab as mlab;
import matplotlib.pyplot as pyplot;
import random;
import math;
import scipy.spatial as sps;

'''
Begin section for defining necessary functions.
'''

# Takes a workbook sheet and a point count and returns an Nx2 numpy array of point values
def recordSheetCoordinates(point_sheet, point_count):
    points = np.zeros((point_count, 3), dtype = np.float);
    for i in range(0, point_count):
        points[i, 0] = point_sheet.cell(row = i + 3, column = 1).value;
        points[i, 1] = point_sheet.cell(row = i + 3, column = 2).value;
    
    # Calculate angle of each point from center (x_mean, y_mean) and sort points by angle 
    x_mean = np.mean(points[:, 0]);
    y_mean = np.mean(points[:, 1]);
    for j in range(0, point_count):
        points[j, 2] = math.atan2(points[j, 1] - y_mean, points[j, 0] - x_mean);
    points = points[points[:,2].argsort()];
    
    # New array with only the x-value and y-value columns
    points2 = np.delete(points,np.s_[-1:],1);
    return points2;

# Check if the given list of points are inside the polygon specified by the list of
# polygon_points. Returns the indices of the points_list whose points are inside the
# given polygon
def checkPointsInPolygon(polygon_list, points_list):
    inside = mlab.inside_poly(points_list, polygon_list);
    return inside;

# Takes the number of random points to be generated, the maximum x-coordinate value
# of the cotyledon outline, and the maximum y-coordinate value of the cotyledon
# outline and returns an Nx2 numpy array of random points.
def generateRandomPoints(number_of_points, x_max, y_max):
    random_points = np.zeros((number_of_points, 2), dtype = np.float);
    for i in range(0, number_of_points):
        random_points[i, 0] = random.uniform(0, x_max);
        random_points[i, 1] = random.uniform(0, y_max); 
    return random_points;

# Takes an Nx2 numpy array of points and returns an array of all the 
# distances representing the autocorrelation of the points
def computeAutocorrelation(points_list):
    return computeCrosscorrelation(points_list, points_list);

# Takes two Nx2 numpy arrays of points and returns an array of all the 
# distances representing the cross correlation of the two point sets
def computeCrosscorrelation(points_list_1, points_list_2):
    point_count_1 = len(points_list_1);
    point_count_2 = len(points_list_2);
    crosscorr_count = point_count_1 * point_count_2;
    crosscorr_list = np.zeros(crosscorr_count, dtype = np.float);
    # Compute each distance and add to list
    for i in range(0, point_count_1):
        for j in range(0, point_count_2):
            crosscorr_list[i*point_count_2 + j] = sps.distance.euclidean(points_list_1[i,:], points_list_2[j,:]);
    return crosscorr_list;

'''
End of function definition section.
'''    
    

'''
Begin of procedural section.
'''

# Reads in the Excel File and gets the worksheet names
# MAKE SURE EXCEL WORKBOOK FILE IS IN FORMAT .xlsx, NOT .xls, OTHERWISE
# THE CODE WILL NOT RUN!
crelox_data_file_name = sys.argv[1];
crelox_wb = openpyxl.load_workbook(crelox_data_file_name);
crelox_wb_sheet_names = crelox_wb.get_sheet_names();

# Reads in the data from the worksheets
stomata_sheet = crelox_wb.get_sheet_by_name('Stomatal Positions');
cotyledon_sheet = crelox_wb.get_sheet_by_name('Cotyledon Outline');

stomata_count = stomata_sheet.max_row - 2;
cotyledon_point_count = cotyledon_sheet.max_row - 2;

stomata_points = recordSheetCoordinates(stomata_sheet, stomata_count);
cotyledon_points = recordSheetCoordinates(cotyledon_sheet, cotyledon_point_count);
cot_x_max = cotyledon_points[:,0].max();
cot_y_max = cotyledon_points[:,1].max();

# Gets the number of sector outline worksheets. Relies on the fact that the sector
# worksheets come sequentially immediately following the cotyledon outline worksheet
sector_names = ['Sector 1 Outline', 'Sector 2 Outline', 'Sector 3 Outline', 'Sector 4 Outline', 'Sector 5 Outline', 'Sector 6 Outline', 'Sector 7 Outline', 'Sector 8 Outline', 'Sector 9 Outline', 'Sector 10 Outline', 'Sector 11 Outline', 'Sector 12 Outline', 'Sector 13 Outline', 'Sector 14 Outline', 'Sector 15 Outline'];
cot_sheet_index = crelox_wb_sheet_names.index('Cotyledon Outline');
number_of_sectors = len(crelox_wb_sheet_names) - (cot_sheet_index + 1);

# Gets the correct worksheets for sector outlines based on number_of_sectors
# Plots points for each sector's outline
for i in range(0, number_of_sectors):
    sector_sheet = crelox_wb.get_sheet_by_name(sector_names[i]);
    sector_point_count = sector_sheet.max_row - 2;
    sector_points = recordSheetCoordinates(sector_sheet, sector_point_count);
    pyplot.plot(sector_points[:,0], sector_points[:,1], 'g');

# Plots the cotyledon outline and stomatal points
pyplot.plot(stomata_points[:,0], stomata_points[:,1], 'b.');
pyplot.plot(cotyledon_points[:,0], cotyledon_points[:,1], 'm');
pyplot.axis('equal');


# Generates 1000 random points with cot_x_max and cot_y_max and check which random
# points are inside the cotyledon
all_rand_points = generateRandomPoints(1000, cot_x_max, cot_y_max);
rand_points_indices_inside = checkPointsInPolygon(cotyledon_points, all_rand_points);

# Create np array with only the random points inside the cotyledon
number_inside = len(rand_points_indices_inside);
rand_points_inside = np.zeros((number_inside, 2), dtype=np.int);
for i in range(0, number_inside):
    rand_points_inside[i] = all_rand_points[rand_points_indices_inside[i]];


#pyplot.plot(all_rand_points[:,0], all_rand_points[:,1], 'r.');
pyplot.plot(rand_points_inside[:,0], rand_points_inside[:,1], 'r.');

pyplot.xlabel('X Position (microns)');
pyplot.ylabel('Y Position (microns)');
pyplot.show();

print("Number of stomata:");
print(stomata_count);
print("Number of random points inside cotyledon:");
print(number_inside);

# Compute autocorrelation of stomatal points and random points
stomata_autocorr = computeAutocorrelation(stomata_points);
random_autocorr = computeAutocorrelation(rand_points_inside);
stom_rand_crosscorr = computeCrosscorrelation(stomata_points, rand_points_inside);


# Plot distributions on histogram
distance_bins = [w * 100 for w in range(25)];
stomata_hist = pyplot.hist(stomata_autocorr, weights=np.ones_like(stomata_autocorr) / stomata_autocorr.size, bins = distance_bins, alpha = 0.5, label = 'Stomata', color = 'b');
random_hist = pyplot.hist(random_autocorr, weights=np.ones_like(random_autocorr) / random_autocorr.size, bins = distance_bins, alpha = 0.5, label = 'Random', color = 'r');
crosscorr_hist = pyplot.hist(stom_rand_crosscorr, weights=np.ones_like(stom_rand_crosscorr) / stom_rand_crosscorr.size, bins = distance_bins, alpha = 0.5, label = 'Crosscorrelation', color = 'w');
pyplot.legend(loc='upper right');
pyplot.xlabel('Distance (microns)');
pyplot.ylabel('Relative Frequency');
pyplot.show();

'''
End of procedural section.
'''